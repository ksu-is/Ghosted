from openpyxl import load_workbook
from db import get_db_connection


def clean_value(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_status(status, interview):
    status = clean_value(status).lower()
    interview = clean_value(interview).lower()

    if status == "rejected":
        return "Rejected"

    if status in ["offer", "offered"]:
        return "Offer"

    if status in ["interviewing", "interview"]:
        return "Interviewing"

    # If status is blank but interview column has a round, count as interviewing
    if interview:
        return "Interviewing"

    return "Applied"


def import_excel_file(file_path):
    workbook = load_workbook(file_path, data_only=True)

    if "All" in workbook.sheetnames:
        sheet = workbook["All"]
    else:
        sheet = workbook.active

    print("IMPORTING SHEET:", sheet.title)
    print("TOTAL EXCEL ROWS:", sheet.max_row)

    headers = [clean_value(cell.value) for cell in sheet[1]]

    imported_count = 0
    skipped_rows = 0

    conn = get_db_connection()

    conn.execute("DELETE FROM applications")
    conn.commit()
    
    seen_rows = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))

        company = clean_value(data.get("Company"))
        role = clean_value(data.get("Position"))
        location = clean_value(data.get("Location"))
        response = clean_value(data.get("Response"))
        interview = clean_value(data.get("Interview"))
        status = normalize_status(data.get("Status"), data.get("Interview"))

        
        if not company and not role:
            skipped_rows += 1
            continue

        row_key = (company, role, location, response, interview, status)

        if row_key in seen_rows:
            skipped_rows += 1
            continue

        seen_rows.add(row_key)

        notes = f"Response: {response}" if response else ""

        offer_status = "Pending"
        if status == "Offer":
            offer_status = "Offer Received"
        elif status == "Rejected":
            offer_status = "Rejected"

        conn.execute(
            """
            INSERT INTO applications
            (company, role, location, date_applied, status, interview_date, follow_up_date, offer_status, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                company,
                role,
                location,
                "",
                status,
                interview,
                "",
                offer_status,
                notes,
            ),
        )

        imported_count += 1

    conn.commit()
    conn.close()

    return imported_count, skipped_rows